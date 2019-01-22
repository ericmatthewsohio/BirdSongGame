# revision .24
import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import random
import re
import os
import pickle
from pygame import mixer
import openpyxl
import collections
import configparser

config = configparser.SafeConfigParser()
config.read('support_files/config.ini')
LARGE_FONT= ("Helvetica", 18)
SMALL_FONT= ("Helvetica", 14)
SMALL_FONT_ul= ("Helvetica", 12, "underline", "bold")
theme_color= (config.get('default', 'theme_color'))
theme_color2= (config.get('default', 'theme_color2'))
theme_color3= (config.get('default', 'theme_color3'))
path_audio = (config.get('default', 'path_audio'))
path_images = (config.get('default', 'path_images'))
path_saves = (config.get('default', 'path_saves'))
dBase = (config.get('default', 'dBase'))
default_image = path_images+"Default.jpg"
bkgd_img = path_images+"bg.gif"
positions = []
recent_misses = []      #only needed in case of missing pickle file


class storage_interface(object):
    passeriformes = []      #create empty list
    non_passeriforms = []   #create empty list
    waterbirds = []         #create empty list
    animal_list = []        #create empty list

    def __init__(self):
        self.wb = openpyxl.load_workbook(dBase)  #open spreadsheet
        self.ws = self.wb['Sheet1']       #define worksheet
        self.ws2 = self.wb['Sheet2']      #define worksheet
        self.ws3 = self.wb['Sheet3']      #define worksheet
        self.ws4 = self.wb['Sheet4']      #define worksheet

    def get_info(self, common_name):
        list = []

        for item in self.ws['D']:
            if item.value == common_name:
                x = (item.row)
                y = (self.ws)
            continue
        for item in self.ws2['D']:
            if item.value == common_name:
                x = (item.row)
                y = (self.ws2)
            continue
        for item in self.ws3['D']:
            if item.value == common_name:
                x = (item.row)
                y = (self.ws3)
            continue
        for item in self.ws4['D']:
            if item.value == common_name:
                x = (item.row)
                y = (self.ws4)
            continue

        for item in y[x]:
            list.append(item.value)
        return list

    def startup_lists(self):
        for item in self.ws['D']:
            if item.value is None:
                continue
            storage_interface.passeriformes.append(item.value)
        for item in self.ws2['D']:
            if item.value is None:
                continue
            storage_interface.non_passeriforms.append(item.value)
        for item in self.ws3['D']:
            if item.value is None:
                continue
            storage_interface.waterbirds.append(item.value)
        for item in self.ws4['D']:
            if item.value is None:
                continue
            storage_interface.animal_list.append(item.value)

c= storage_interface()
c.startup_lists()
mixer.init()

def stop_song():
    mixer.music.fadeout(300) 

def replay_song(event=None):
    mixer.music.play(0)

def save_misses_file():                                 #saves 100 misses to file 
    with open('support_files/recent_misses.pickle', 'wb') as fp:      #save selections to file
        pickle.dump(recent_misses, fp)
    

def load_misses_file():                                 #loads 100 misses from file
    try:
        with open ('support_files/recent_misses.pickle', 'rb') as fp:     #load file with miss history
            misses = pickle.load(fp)
            #recent_misses.append(misses)
            return misses
    except Exception:
        pass
misses = load_misses_file()
try:
    recent_misses = collections.deque(misses,maxlen=100)    # creates que of last 100 misses
except Exception:
    pass

class BirdSong(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        
        #tk.Tk.iconbitmap(self,default='English.ico')
        tk.Tk.wm_title(self, "Bird Song")
        tk.Tk.geometry(self, "800x768+600+230") #this format does the same as the next line
        #self.geometry("1024x768+230+230") ((same as above line))
        self.bind('<Escape>', lambda e: self.shut_down())
        self.bind()
        self.protocol("WM_DELETE_WINDOW", lambda: self.shut_down())
        self.resizable(False, False)
        self.tk_setPalette(background=theme_color, selectBackground=theme_color3,
               activeBackground='black', activeForeground=theme_color3, troughcolor="red")
        s = ttk.Style()
        s.theme_use("default")
        s.configure('TButton', font=LARGE_FONT, background=theme_color)
        s.map('TButton',
            background=[('pressed', 'focus', theme_color3),
            ('active', theme_color)])
        s.configure('my.TButton', font=LARGE_FONT, background=theme_color)
        
        s.configure('TLabelframe', background=theme_color)
        s.configure('TLabel', font=LARGE_FONT, background=theme_color)  
        s.configure('TRadiobutton', background=theme_color)
        s.configure('TFrame', background=theme_color)
        s.configure('TScrollbar', background=theme_color, troughcolor=theme_color2)
        s.configure('TNotebook', background=theme_color)
        s.configure("TNotebook.Tab", background=theme_color)
        s.map("TNotebook.Tab", 
            background=[("selected", theme_color2)])
        s.map('TScrollbar',
            foreground=[('disabled', 'yellow'),
            ('pressed', 'red'),
            ('active', 'blue')],
            background=[('disabled', theme_color),
            ('pressed', '!focus', theme_color3),
            ('active', theme_color2)],)
        container = ttk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (StartPage, GamePage, ConfigPage, HelpPage):

            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)
    
    def show_frame(self, page_name):
        frame = self.frames[page_name]
        #print(page_name)
        #print(frame)
        frame.tkraise()
        menubar = frame.menubar(self)   ##menubar
        self.configure(menu=menubar)    ##menubar
    
    def get_page(self, classname):
        '''Returns an instance of a page given it's class name as a string'''
        for page in self.frames.values():
            if str(page.__class__.__name__) == classname:
                return page
        return None
    def shut_down(self):                                        # controlled shutdown command
        print('shutting down and saving recent misses')
        c = save_misses_file()
        self.after(1000, self.destroy())

class StartPage(ttk.Frame):
    def menubar(self, root):        ##menubar
        menubar = tk.Menu(root)
        return menubar 
    def __init__(self, parent, controller):
        self.controller = controller
        ttk.Frame.__init__(self,parent)
        self.subframe = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe.pack(fill='x')
        self.subframe1 = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe1.pack(fill='both', expand=True)
        self.subframe1.columnconfigure(0, weight=1, minsize = 100)
        self.subframe1.columnconfigure(1, weight=1, minsize = 100)
        self.subframe1.rowconfigure(0, weight=1, minsize = 100)
        self.subframe1.rowconfigure(1, weight=1, minsize = 100)
        self.load_widgets(controller)
        
    def load_widgets(self, controller):
        self.img = tk.PhotoImage(file=path_images+"Default3.gif")
        label = ttk.Label(self.subframe, text="Learn Bird Songs and More", font=LARGE_FONT)
        self.label=tk.Label(self.subframe1, image = self.img)

        #self.game_tile_image= tk.PhotoImage(file=path_images+'bcc2.gif')
        #self.game_tile = ttk.Button(self.subframe1,  image=self.game_tile_image, text ="Play Game", compound = tk.TOP,
        #                    command=lambda: controller.show_frame(GamePage))
        #self.help_tile_image= tk.PhotoImage(file=path_images+'bird.png')
        #self.help_tile = ttk.Button(self.subframe1,  image=self.help_tile_image, text ="Help & Info", compound = tk.TOP,
        #                    command=lambda: controller.show_frame(HelpPage))
        #self.config_tile_image= tk.PhotoImage(file=path_images+'bird.png')
        #self.config_tile = ttk.Button(self.subframe1,  image=self.config_tile_image, text ="Configure Options", compound = tk.TOP,
        #                    command=lambda: controller.show_frame(ConfigPage))

        #self.game_tile.grid(column=0, row=0, columnspan = 2)
        #self.help_tile.grid(column=0, row=1)
        #self.config_tile.grid(column=1, row=1)
        #label.pack(pady=10,padx=10)
        self.label.place(relwidth=1, relheight=1)
        self.label.bind('<1>', self.click)
        self.label.bind('<Motion>', self.motion)

    def click(self,event):
        x, y = event.x, event.y
        print('{}, {}'.format(x, y))
        if (x>325 and x<460) and (y>165 and y<310):
            self.controller.show_frame(GamePage)
        elif (x>180 and x<300) and (y>450 and y<600):
            x=self.controller.show_frame(HelpPage)
        elif (x>475 and x<600) and (y>450 and y<600):
            self.controller.show_frame(ConfigPage)

    def motion(self,event=None):
        x, y = event.x, event.y
        #print('{}, {}'.format(x, y))
        if (x>325 and x<460) and (y>165 and y<310):
            self.label.configure(cursor = "hand2")
        elif (x>180 and x<300) and (y>450 and y<600):
            self.label.configure(cursor = "hand2")
        elif (x>475 and x<600) and (y>450 and y<600):
            self.label.configure(cursor = "hand2")
        else:
            self.label.configure(cursor="")

class GamePage(ttk.Frame):
    number_correct = 0
    number_wrong = 0
    image_showing = path_images+"Default.jpg"
    my_list2 = ()
    bird_name ='NorthernCardinal'
    def menubar(self, root):        ##menubar
        menubar = tk.Menu(root)
        return menubar 

    def __init__(self, parent, controller):
        self.controller = controller
        ttk.Frame.__init__(self, parent)
        self.subframe = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe.pack(fill='x')
        self.l = tk.Label(self,text = 'Game Stats', background = theme_color)
        self.labelframe = ttk.LabelFrame(self.subframe, text="Game Stats",labelwidget = self.l)
        
        self.labelframe.pack(side = "right", padx=5, pady=1)
        self.subframe1 = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe1.pack(fill='both', expand=True)
        self.subframe2 = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe2.pack(fill='x')
        self.load_widgets()

    def load_widgets(self):
        self.img = tk.PhotoImage(file=bkgd_img)
        self.config_icon= tk.PhotoImage(file=path_images+'gear.png')
        self.home_icon= tk.PhotoImage(file=path_images+'home.png')
        self.info_icon= tk.PhotoImage(file=path_images+'info.png')

        label1=tk.Label(self.subframe, image = self.img)
        label2=tk.Label(self.subframe1, image = self.img)
        label3=tk.Label(self.subframe2, image = self.img)
        button1 = tk.Button(self.subframe, image=self.home_icon, cursor = "hand2", command=lambda: GamePage.home(self))
        button01 = tk.Button(self.subframe, image=self.config_icon, cursor = "hand2", command=lambda: GamePage.config(self))
        button02 = tk.Button(self.subframe, image=self.info_icon,font = LARGE_FONT, cursor = "hand2", command=lambda: GamePage.help(self))
        self.label = tk.Label(self.subframe, text="Click Start to Begin",width=26,borderwidth=1,
                              relief='raised', font=LARGE_FONT,  anchor='center')
        self.labela = tk.Label(self.subframe, text="",width=26,borderwidth=1,
                              relief='raised', font=("Helvetica", 14, "italic"), anchor='center')
        self.label10 = tk.Label(self.labelframe, text="100%",font = ("Helvetica", 16), width=6, anchor='center')
        self.label11 = tk.Label(self.labelframe, text="0"+" ⬆",font = ("Helvetica", 10), width=6, anchor='center')
        self.label12 = tk.Label(self.labelframe, text="0"+" ⬇",font = ("Helvetica", 10), width=6, anchor='center')

        self.label2 = tk.Label(self.subframe1)
        self.image_show(default_image)

        self.button2 = tk.Button(self.subframe2,font = LARGE_FONT, text="replay song",command= replay_song)
        self.button3 = tk.Button(self.subframe2,font = LARGE_FONT, text="Start",width=8,command=lambda: self.start())
        self.button4 = tk.Button(self.subframe2,font = LARGE_FONT, text="Hidden",width=8,command=lambda: self.not_sure())

        label1.place(relwidth=1, relheight=1)
        label2.place(relwidth=1, relheight=1)
        label3.place(relwidth=1, relheight=1)
        button1.pack(side='left', padx=[4,2], pady=6)
        button01.pack(side='left', padx=2, pady=1)
        button02.pack(side='left', padx=2, pady=1)
        self.label.place(relx=0.5, rely=0.25, anchor='center')
        self.labela.place(relx=0.5, rely=0.75, anchor='center')
        self.label10.pack(side = "left", padx=6, pady=1)
        self.label11.pack(side = "top", padx=6, pady=1)
        self.label12.pack(side = "bottom", padx=6, pady=1)
        self.label2.pack(side = "bottom", expand = "true")
        #button2.pack(side='left', padx=10, pady=10)
        self.button3.pack(side='right', padx=5, pady=5)
        #button4.pack(side='right', padx=5, pady=10)
        self.labelframe.lift()

    def home(self):
        stop_song()       
        self.button3.configure(text = "Start", command=lambda: self.start())
        self.button4.pack_forget()
        self.image_show(default_image)
        self.label.configure(text = "   Game Page   ")
        self.labela.configure(text = "     ")
        self.controller.show_frame(StartPage)


    def config(self):
        stop_song()       
        self.button3.configure(text = "Start", command=lambda: self.start())
        self.button4.pack_forget()
        self.image_show(default_image)
        self.label.configure(text = "   Game Page   ")
        self.labela.configure(text = "      ")
        self.controller.show_frame(ConfigPage)

    def help(self):
        stop_song()       
        self.button3.configure(text = "Start", command=lambda: self.start())
        self.button4.pack_forget()
        self.image_show(default_image)
        self.label.configure(text = "   Game Page   ")
        self.labela.configure(text = "      ")
        self.controller.show_frame(HelpPage)

    def start(self, event=None):
        self.label.configure(text = "")
        self.labela.configure(text = "")
        self.labela.place(relx=0.5, rely=0.75, anchor='center')
        self.button3.configure(text = "Know", command=lambda: self.know())
        self.button4.configure(text = "Not Sure", command=lambda: self.not_sure())
        self.button4.pack(side='right', padx=5, pady=5)
        self.button2.pack(side='left', padx=10, pady=5)
        page_two = self.controller.get_page("ConfigPage")
        value = page_two.game_list
        self.bird_name = self.random_species(value)
        self.subframe2.focus_set()
        self.subframe2.bind("<KeyPress-Right>", self.know)
        self.subframe2.bind("<KeyPress-Left>", self.not_sure)
        self.subframe2.bind("<KeyPress-Down>", replay_song)
        page_two = self.controller.get_page("ConfigPage")
        value = page_two.var.get()
        if str(value) == "1":
            self.play_random_song(self.bird_name)
            self.image_show(default_image)
        elif str(value) == "3":
            self.play_random_song(self.bird_name)
            self.playingsong_showimage()
        else: self.playingsong_showimage()
    
    def know(self, event=None):
        stop_song()
        self.button3.configure(text = "Got It", command=lambda: self.got_it())
        self.button4.configure(text = "Missed It", command=lambda: self.missed_it())
        p = re.compile(r'([a-z])([A-Z])') # add space before capital letter in bird_name
        add_space = re.sub(p, r"\1 \2", self.bird_name) # add space before capital letter in bird_name
        self.label.configure(text = add_space)
        
        c= storage_interface()
        d=c.get_info(self.bird_name)
        self.labela.configure(text = d[8])
        
        page_two = self.controller.get_page("ConfigPage")
        value = page_two.var.get()
        if str(value) == "1":
            self.playingsong_showimage()
        self.subframe2.focus_set()
        self.subframe2.bind("<KeyPress-Right>", self.got_it)
        self.subframe2.bind("<KeyPress-Left>", self.missed_it)
        self.subframe2.bind("<KeyPress-Down>", replay_song)
    
    def got_it(self, event=None):
        self.start()
        self.number_correct += 1
        self.label10.configure(text="%4.0f"%((self.number_correct/(self.number_correct+self.number_wrong))*100)+"%")
        self.label11.configure(text=str(int(self.number_correct))+" ⬆")
        
    def hidden(self):
        self.button4.configure(text = "Dont Know", command=lambda: self.not_sure())
            
    def not_sure(self, event=None):
        stop_song()
        self.button4.configure(text = "next", command=lambda: self.start())
        self.button3.configure(text = "Next", command=lambda: self.start())
        self.button4.pack_forget()
        self.subframe2.focus_set()
        self.subframe2.bind("<KeyPress-Right>", self.start)
        self.subframe2.bind("<KeyPress-Down>", replay_song)
        p = re.compile(r'([a-z])([A-Z])') # add space before capital letter in bird_name
        add_space = re.sub(p, r"\1 \2", self.bird_name) # add space before capital letter in bird_name
        self.label.configure(text = add_space)
        c= storage_interface()
        d=c.get_info(self.bird_name)
        self.labela.configure(text = d[8])
        self.labela.place(relx=0.5, rely=0.75, anchor='center')
        page_two = self.controller.get_page("ConfigPage")
        value = page_two.var.get()
        if str(value) == "1":
            self.playingsong_showimage()
            
    def missed_it(self, event=None):
        self.number_wrong += 1
        self.label10.configure(text="%4.0f"%((self.number_correct/(self.number_correct+self.number_wrong))*100)+"%")
        self.label12.configure(text=str(int(self.number_wrong))+" ⬇")
        recent_misses.append(self.bird_name)
        #most_common = collections.Counter(recent_misses).most_common(5)
        #self.my_list2, my_list1 = zip(*most_common)     #my_list2 is birds, my_list1 is the number of misses
        #print(self.my_list2)
        self.start()
     
    def playingsong_showimage(self):
        imagebb = path_images +(self.bird_name)+".jpg"
        self.image_show(imagebb)
        
    def image_show(self, img):
        self.count = 0
        self.image_new = img
        self.image = Image.open(self.image_new)                                    # load source image
        self.image.thumbnail((755, 596))                                # shrink source image and force proper dimensions
        self.imgSize = self.image.size                                  # find dimensions of image
        self.new_im = Image.new("RGB", (755, 596), color = theme_color)
        self.new_im.paste(self.image, ((755-self.imgSize[0])//2,
                    (596-self.imgSize[1])//2))

        self.image_two = self.image_showing
        self.image_two = Image.open(self.image_showing)                                    # load source image
        self.image_two.thumbnail((755, 596))                                # shrink source image and force proper dimensions
        self.imgSize2 = self.image_two.size
        self.new_im2 = Image.new("RGB", (755, 596), color = theme_color)
        self.new_im2.paste(self.image_two, ((755-self.imgSize2[0])//2,
                    (596-self.imgSize2[1])//2))
        
        self.fade_label2()

    def fade_label2(self):
        try:
            if self.count < 1:
                self.blend=Image.blend(self.new_im2, self.new_im, alpha = self.count)
                self.rendered_img = ImageTk.PhotoImage(self.blend)           # convert image for Tk compatibility
                self.label2.configure(image = self.rendered_img)             # apply image to label
                self.after(1, self.fade_label2)                         # call this method again in 1 millisecond
                self.count += .07                                       # fade by x steps
            else:
                
                self.image_showing = self.image_new
        except FileNotFoundError:
            tk.messagebox.showerror(title="Show Error", message = "Oops! " + self.bird_name +" missing image file")


    def random_species(self, game):                                       # return random species to quiz
        species_to_quiz = (random.choice(game))
        #print(species_to_quiz)
        return species_to_quiz

    def play_random_song(self, species_to_quiz):                      # create list of all species songs and calls
        songs = []
        page_two = self.controller.get_page("ConfigPage")
        value = page_two.var2.get()
        for file in os.listdir(path_audio):
            filename = os.fsdecode(file)
            if filename.startswith(species_to_quiz): 
                songs.append(filename)

        if value == 1:                                     # filters to one song or call for easy difficulty
            ssongs = []
            for s in songs:
                if s[-5] in '1':
                    ssongs.append(s)
        elif value == 2:                                   # filters to three songs and/or calls for intermediate difficulty
            ssongs = []
            for s in songs:
                if s[-5] in '123':
                    ssongs.append(s)
        else: ssongs = songs                                    # passes ass songs and/or calls for maximum difficulty
        #print(ssongs)
        try:
            song_to_quiz = (random.choice(ssongs))
            mixer.music.load(path_audio +(song_to_quiz))
            mixer.music.play(0)
            a = mixer.Sound(path_audio +(song_to_quiz))
            #print("length of song is:",int(a.get_length()))    # return length of song in seconds
            #print("filename is: ", song_to_quiz)           # return filename of song
        except IndexError: 
            tk.messagebox.showerror(title="Show Error", message = "Oops! " + species_to_quiz +" audio files are missing")


class ConfigPage(ttk.Frame):
    def menubar(self, root):        ##menubar
        controller = self.controller
        menubar = tk.Menu(root)
        fileMenu = tk.Menu(menubar)
        fileMenu.add_command(label="Open playlist", command = self.load_file)
        fileMenu.add_command(label="Save current playlist", command =lambda: ConfigPage.save_file(self, controller))
        fileMenu.add_command(label="Create 'most missed' playlist", command =lambda: ConfigPage.save_file_most_missed(self, controller))
        menubar.add_cascade(label="Saved Playists", menu=fileMenu)
        return menubar 

    def find_posns_in_listbox(user_list, master_list):
        positions = []
        for i in user_list:
            try:
                index_element = master_list.index(i)
                positions.append(index_element)
            except ValueError:
                continue
        return positions

    def __init__(self, parent, controller):
        self.controller = controller
        tk.Frame.__init__(self, parent)

        self.subframe = tk.Frame(self, relief='raised', borderwidth=3)
        self.subframe.pack(fill='x')
        
        
        self.subframe1 = tk.Frame(self, relief='raised', borderwidth=3)
        self.subframe1.pack(fill='both', expand=True)
        self.subframe2 = tk.Frame(self, relief='raised', borderwidth=3)
        self.subframe2.pack(fill='x')
        self.l = tk.Label(self,text = 'Game Type', background = theme_color)
        self.labelframe1 = tk.LabelFrame(self.subframe1, text="Game Type", labelwidget = self.l)
        self.labelframe1.grid(row = 7, rowspan = 3, column = 0, sticky='W')
        self.l2 = tk.Label(self,text = 'Game Difficulty', background = theme_color)
        self.labelframe2 = tk.LabelFrame(self.subframe1, text="Game Difficulty", labelwidget = self.l2)
        self.labelframe2.grid(row = 7, rowspan = 3, column = 2, sticky='W')
        self.load_widgets()

    def load_widgets(self):
        self.img = tk.PhotoImage(file=bkgd_img)
        self.home_icon= tk.PhotoImage(file=path_images+'home.png')
        self.game_icon= tk.PhotoImage(file=path_images+'bird_icon.png')
        self.info_icon= tk.PhotoImage(file=path_images+'info.png')
        
        label1=tk.Label(self.subframe, image = self.img)
        label2=tk.Label(self.subframe1, image = self.img)
        label3=tk.Label(self.subframe2, image = self.img)
        label = tk.Label(self.subframe, text="Configuration", font=LARGE_FONT)
        button1 = tk.Button(self.subframe, image=self.home_icon, cursor = "hand2", command=lambda: ConfigPage.st_page(self))
        button3 = tk.Button(self.subframe, image=self.info_icon,cursor = "hand2", command=lambda: ConfigPage.help_page(self))
        button2 = tk.Button(self.subframe, image=self.game_icon,cursor = "hand2", command=lambda: ConfigPage.game_page(self))
        

        self.var = tk.IntVar(value=1)
        var = self.var
        R1 = tk.Radiobutton(self.labelframe1, width = 15, anchor = tk.W,  text="Sound Only", variable=self.var, value=1,command=None)
        R2 = tk.Radiobutton(self.labelframe1, anchor = tk.W,  text="Image Only", variable=self.var, value=2,command=None)
        R3 = tk.Radiobutton(self.labelframe1, anchor = tk.W,  text="Sound and Image", variable=self.var, value=3,command=None)

        self.var2 = tk.IntVar(value=3)
        var2 = self.var2
        R4 = tk.Radiobutton(self.labelframe2, width = 15, anchor = tk.W,  text="easy", variable=self.var2, value=1,command=None)
        R5 = tk.Radiobutton(self.labelframe2, anchor = tk.W,  text="Intermediate", variable=self.var2, value=2,command=None)
        R6 = tk.Radiobutton(self.labelframe2, anchor = tk.W,  text="advanced", variable=self.var2, value=3,command=None)
        
        quiz_test_list = []            # do not delete. gives empty set on missing default file start.
        self.listbox1 = tk.Listbox(self.subframe1, selectmode= "extended", width=25, height=25,exportselection=0)
        yscroll = ttk.Scrollbar(self.subframe1, command=self.listbox1.yview, orient=tk.VERTICAL)
        self.listbox1.configure(yscrollcommand=yscroll.set)
        for item in storage_interface.passeriformes:   #load list of birds to listbox
            self.listbox1.insert(tk.END, item)
        self.listbox2 = tk.Listbox(self.subframe1, selectmode= "extended", width=25, height=25,exportselection=0)
        
        yscroll2 = ttk.Scrollbar(self.subframe1, command=self.listbox2.yview, orient=tk.VERTICAL)
        self.listbox2.configure(yscrollcommand=yscroll2.set)
        for item in storage_interface.non_passeriforms:   #load list of animals to listbox
            self.listbox2.insert(tk.END, item)
        self.listbox3 = tk.Listbox(self.subframe1, selectmode= "extended", width=25, height=25,exportselection=0)
        yscroll3 = ttk.Scrollbar(self.subframe1, command=self.listbox3.yview, orient=tk.VERTICAL)
        self.listbox3.configure(yscrollcommand=yscroll3.set)
        for item in storage_interface.waterbirds:   #load list of animals to listbox
            self.listbox3.insert(tk.END, item)
        self.listbox4 = tk.Listbox(self.subframe1, selectmode= "extended", width=25, height=25,exportselection=0)
        yscroll4 = ttk.Scrollbar(self.subframe1, command=self.listbox4.yview, orient=tk.VERTICAL)
        self.listbox4.configure(yscrollcommand=yscroll4.set)
        for item in storage_interface.animal_list:   #load list of animals to listbox
            self.listbox4.insert(tk.END, item)

        btn1 = tk.Label(self.subframe1, pady=(5), font = SMALL_FONT_ul, text="passeriformes")
        btn2 = tk.Label(self.subframe1, pady=(5), font = SMALL_FONT_ul, text="non-passeriformes")
        btn3 = tk.Label(self.subframe1, pady=(5), font = SMALL_FONT_ul, text="waterbirds" )
        btn4 = tk.Label(self.subframe1, pady=(5), font = SMALL_FONT_ul, text="other sounds")

        
        self.active_selections = tk.Label(self.subframe1, font = SMALL_FONT, text = "Active selections:")
        

        label1.place(relwidth=1, relheight=1)
        label2.place(relwidth=1, relheight=1)
        label3.place(relwidth=1, relheight=1)
        button1.pack(side='left', padx=[4,2], pady=5)
        button3.pack(side='left', padx=2, pady=1)
        button2.pack(side='left', padx=2, pady=1)
        label.place(relx=0.5, rely=0.5, anchor='center')
        R1.grid(row=0, column=0, sticky=tk.N+tk.S+tk.W+tk.E)
        R2.grid(row=1, column=0, sticky=tk.N+tk.S+tk.W+tk.E)
        R3.grid(row=2, column=0, sticky=tk.N+tk.S+tk.W+tk.E)
        R4.grid(row=0, column=0, sticky=tk.N+tk.S+tk.W+tk.E)
        R5.grid(row=1, column=0, sticky=tk.N+tk.S+tk.W+tk.E)
        R6.grid(row=2, column=0, sticky=tk.N+tk.S+tk.W+tk.E)
        self.listbox1.grid(row=1, column=0)
        yscroll.grid(row=1, column=1, sticky=tk.N+tk.S)
        self.listbox2.grid(row=1, column=2)
        yscroll2.grid(row=1, column=3, sticky=tk.N+tk.S)
        self.listbox3.grid(row=1, column=4)
        yscroll3.grid(row=1, column=5, sticky=tk.N+tk.S)
        self.listbox4.grid(row=1, column=6)
        yscroll4.grid(row=1, column=7, sticky=tk.N+tk.S)
        btn1.grid(column=0, row=0)
        btn2.grid(column=2, row=0)
        btn3.grid(column=4, row=0)
        btn4.grid(column=6, row=0)
        self.labelframe1.lift()
        self.labelframe2.lift()
        self.active_selections.grid(row=10, column=0, columnspan=3, sticky='W')

        ConfigPage.populate_listbox(self, 'support_files/default.pickle')
        ConfigPage.activate_game_list(self)

    def st_page(self):
        ConfigPage.activate_game_list(self)
        self.controller.show_frame(StartPage)

    def game_page(self):
        ConfigPage.activate_game_list(self)
        self.controller.show_frame(GamePage)

    def help_page(self):
        ConfigPage.activate_game_list(self)
        self.controller.show_frame(HelpPage)
        
    def populate_listbox(self, file_name):
        quiz_test_list = []
        sticky_defaults = []
        try:
            with open (file_name, 'rb') as fp:  #open file with selection from last played game
                quiz_test_list = pickle.load(fp)    #loads list of quiz items
                sticky_defaults = pickle.load(fp)   #sets game type and difficulty
                #print(sticky_defaults)
                self.var.set(sticky_defaults[0])
                self.var2.set(sticky_defaults[1])
        except FileNotFoundError:
            tk.messagebox.showerror(title="Show Error", message = "Oops! Default file missing. Creating new one. Go to configuration page to create/activate a playlist.")
            print("File Does Not Exist")
        except Exception as e:
            tk.messagebox.showerror(title="Show Error", message = "Oops! Invalid file. Choose a valid file type.")
            print (e)
            return
        bob = ConfigPage.find_posns_in_listbox(quiz_test_list, storage_interface.passeriformes)    
        self.listbox1.selection_clear(0, 'end')
        for item in bob:  #set selected items in list from file
            self.listbox1.select_set(item)    
        bob1 = ConfigPage.find_posns_in_listbox(quiz_test_list, storage_interface.non_passeriforms)       
        self.listbox2.selection_clear(0, 'end')
        for item in bob1:
            self.listbox2.select_set(item)    
        bob2 = ConfigPage.find_posns_in_listbox(quiz_test_list, storage_interface.waterbirds)        
        self.listbox3.selection_clear(0, 'end')
        for item in bob2:
            self.listbox3.select_set(item)    
        bob3 = ConfigPage.find_posns_in_listbox(quiz_test_list, storage_interface.animal_list)
        self.listbox4.selection_clear(0, 'end')
        for item in bob3:
            self.listbox4.select_set(item)
        self.active_selections.configure(text = "Active selections:"+ str(len(quiz_test_list)))
     
    def activate_game_list(self): 
        self.game_list = []
        self.game_defaults = [self.var.get(), self.var2.get()]
        selection = self.listbox1.curselection()
        selection2 = self.listbox2.curselection()
        selection3 = self.listbox3.curselection()
        selection4 = self.listbox4.curselection()
        for i in selection:
            entrada = self.listbox1.get(i)
            self.game_list.append(entrada)
        for i in selection2:
            entrada = self.listbox2.get(i)
            self.game_list.append(entrada)
        for i in selection3:
            entrada = self.listbox3.get(i)
            self.game_list.append(entrada)
        for i in selection4:
            entrada = self.listbox4.get(i)
            self.game_list.append(entrada)
        #print(selection)
        #print(selection2)
        #print(selection3)
        #print(selection4)
        self.active_selections.configure(text = "Active selections:"+ str(len(self.game_list)))
        with open('support_files/default.pickle', 'wb') as fp:  #save selections to default file
            pickle.dump(self.game_list, fp)
            pickle.dump(self.game_defaults, fp)

    def load_file(self):
        answer = tk.filedialog.askopenfilename(initialdir = path_saves, filetypes=[('text files', 'bird')])
        if answer == (""):
            answer = ('support_files/default.pickle')
        ConfigPage.populate_listbox(self, answer)
        ConfigPage.activate_game_list(self)
        

    def save_file(self, controller):
        ConfigPage.activate_game_list(self)
        answer = tk.filedialog.asksaveasfilename(initialdir = path_saves ,defaultextension = '.bird', filetypes=[('text files', 'bird')])
        with open(answer, 'wb') as fp:  #save current selections to named file
            pickle.dump(self.game_list, fp)
            pickle.dump(self.game_defaults, fp)

    def save_file_most_missed(self, controller):
        most_common = collections.Counter(recent_misses).most_common(10)
        my_list2, my_list1 = zip(*most_common)
        with open(path_saves+'most_missed.bird', 'wb') as fp:  #save most missed to file
            pickle.dump(my_list2, fp)
            pickle.dump(self.game_defaults, fp)

class HelpPage(ttk.Frame):
    def menubar(self, root):        ##menubar
        menubar = tk.Menu(root)
        return menubar 
    def __init__(self, parent, controller):
        self.controller = controller
        ttk.Frame.__init__(self,parent)
        self.subframe = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe.pack(fill='x')
        self.subframe1 = ttk.Frame(self, relief='raised', borderwidth=3)
        self.subframe1.pack(fill='both', expand=True)
        self.load_widgets(controller)
    def load_widgets(self, controller):
        self.nb = ttk.Notebook(self.subframe1)
        self.f1 = ttk.Frame(self.nb)
        self.nb.add(self.f1, text="Gameplay")
        self.f2 = ttk.Frame(self.nb)
        self.nb.add(self.f2, text="Most missed")
        self.f3 = ttk.Frame(self.nb)
        self.nb.add(self.f3, text="Species Info")

        self.img = tk.PhotoImage(file=bkgd_img)
        self.home_icon= tk.PhotoImage(file=path_images+'home.png')
        self.game_icon= tk.PhotoImage(file=path_images+'bird_icon.png')
        self.config_icon= tk.PhotoImage(file=path_images+'gear.png')
        label1=tk.Label(self.subframe, image = self.img)
        label2=tk.Label(self.f1, image = self.img)
        label3=tk.Label(self.f2, image = self.img)
        label4=tk.Label(self.f3, image = self.img)

        button = tk.Button(self.subframe, image=self.game_icon,
                            command=lambda: controller.show_frame(GamePage))
        button1 = tk.Button(self.subframe, image=self.home_icon,
                        command=lambda: controller.show_frame(StartPage))
        button2 = tk.Button(self.subframe, image=self.config_icon,
                            command= lambda: controller.show_frame(ConfigPage))

        ############# Tab 1 ###################
        label = ttk.Label(self.subframe, text="Help & Information", font=LARGE_FONT)
        text_help = scrolledtext.ScrolledText(self.f1, height=32,wrap=tk.WORD, width=65,background = theme_color, font = ("Times New Roman", 14))
        text_help.pack(side='top', fill='y')
        
        quote = """HOW TO PLAY:
Click Play Game to go to the game page, then click Start when ready. 
The Game plays bird sounds and/or shows images, and you try to identify them \
as if you're out birding. Click Know it if you do, and Not Sure, otherwise. \
The Game then shows you the bird and you confirm whether you were correct or not.

CONFIGURATION OPTIONS:

GAME TYPE:
SOUND: The sound is played first, this is the most challenging game type.

IMAGE: Use this to learn birds by sight. Sound is not played.

SOUND and IMAGE: The sound is played, while the image is shown. This is a good mode when ititially learning new birds.

GAME DIFFICULTY:
Easy: Only the most representitive version of the bird call and-or song will be available during gameplay.

Intermediate: Three versions of the bird calls and-or songs will be available during gameplay.

Advanced: All versions of songs and calls will be available during gameplay.
"""
        text_help.insert('end', quote)
        text_help.tag_add('highlightline', '1.0', '1.end')
        text_help.tag_add('tab', '2.0', '3.end')
        text_help.tag_add('highlightline', '5.0', '5.end')
        text_help.tag_add('tab', '8.0', '12.end')
        text_help.tag_add('underline', '7.0', '7.end')
        text_help.tag_add('underline', '14.0', '14.end')
        text_help.tag_add('tab', '15.0', '19.end')
        text_help.tag_configure('highlightline',underline = True, background='yellow', font='helvetica 14 bold', relief='raised')
        text_help.tag_configure('underline',underline = True ,font='helvetica 14 bold', relief='raised')
        text_help.tag_configure('tab',lmargin1 = 50, lmargin2 = 50)
        text_help.configure(state="disabled")
        ############# Tab 2 ###################
        text_missed = scrolledtext.ScrolledText(self.f2, height=32,padx = 15,wrap=tk.WORD, width=65,background = theme_color, font = ("Times New Roman", 14))
        text_missed.pack(side='top', fill='y')
        
        quote_missed = """Most Missed Out of Last 100 Misses:\n"""
        text_missed.insert('end', quote_missed)
        most_common = collections.Counter(recent_misses).most_common(10)
        self.my_list2, my_list1 = zip(*most_common)
        p = re.compile(r'([a-z])([A-Z])') # add space before capital letter in bird_name
        for i in range(0,10):
            add_space = re.sub(p, r"\1 \2", self.my_list2[i]) # add space before capital letter in bird_name
            text_missed.insert('end',str(my_list1[i])+' misses \t\t'+add_space+ '\n')
        text_missed.configure(state="disabled")
        ############# Tab 3 ###################
        self.bird_entry=tk.Label(self.f3,text = "Choose Bird:", width=25)
        self.bird_list=ttk.Combobox(self.f3, width=28)
        self.bird_list.configure(values = storage_interface.passeriformes+storage_interface.non_passeriforms+storage_interface.waterbirds+storage_interface.animal_list)
        self.song_entry=tk.Label(self.f3,text = "Choose Song:", width=25)
        self.song = ttk.Combobox(self.f3, width=28)
        self.text_bird_entry=scrolledtext.ScrolledText(self.f3,height=32,spacing1=3,spacing3=3,padx = 15,wrap=tk.WORD, width=65,background = theme_color, font = ("Times New Roman", 14))
        self.text_bird_entry.configure(state="disabled")
        
        self.bird_entry.pack()
        self.bird_list.pack()
        self.song_entry.pack()
        self.song.pack()
        self.text_bird_entry.pack(side='top', fill='y')

        label1.place(relwidth=1, relheight=1)
        label2.place(relwidth=1, relheight=1)
        label3.place(relwidth=1, relheight=1)
        label4.place(relwidth=1, relheight=1)
        self.nb.pack(expand=1, fill="both")
        label.place(relx=0.5, rely=0.5, anchor='center')
        button1.pack(side='left', padx=5, pady=5)
        button2.pack(side='left', padx=1, pady=1)
        button.pack(side='left', padx=5, pady=5)
        self.f3.bind('<Map>',self.get_species_info)
        self.f3.bind('<Expose>',self.get_species_info)
        self.bird_entry.bind('<Return>',self.enter)
        self.song.bind("<<ComboboxSelected>>", self.play_song)
        self.bird_list.bind("<<ComboboxSelected>>", self.enter2)

    def enter(self, event=None):
        self.get_species_info(bird = self.bird_entry.get())

    def enter2(self, event=None):
        self.get_species_info(bird = self.bird_list.get())

    def get_species_info(self,event=None,bird=None):
        self.text_bird_entry.configure(state="normal")
        page_two = self.controller.get_page("GamePage")
        value = page_two.bird_name
        #print("bird: "+str(bird))
        if bird !=None:
            value = bird
            bird = None
        A = storage_interface()
        species_info = A.get_info(value)

        
        self.text_bird_entry.delete('1.0', 'end')
        
        self.image7 = Image.open(path_images +(str(value))+".jpg")
        self.image7.thumbnail((755/2, 596/2))
        self.rendered_img = ImageTk.PhotoImage(self.image7)
        self.text_bird_entry.insert('end','Species Information Page\n')
        self.text_bird_entry.image_create('end',image=self.rendered_img)
        quote_bird_entry = ("\nCommon name: "+species_info[1]
                            +"\nScientific Name: "+species_info[8]
                            +"\nFamily: "+species_info[5]
                            +"\nOrder: "+species_info[4]
                            )
        self.text_bird_entry.insert('end',quote_bird_entry)
        try:
            self.text_bird_entry.insert('end',"\nVocalizations: "+species_info[9])
        except:
            pass
        self.text_bird_entry.tag_add('underline', '1.0', '1.end')
        self.text_bird_entry.tag_config('underline',justify='center',underline = True ,font='helvetica 14 bold', relief='raised')
        self.text_bird_entry.configure(state="disabled")
        songs = self.song_list(bird = value)
        self.song.configure(values = songs)
        return species_info

    def song_list(self, bird=None):                      # create list of all species songs and calls
        #print(bird)
        songs = []
        page_two = self.controller.get_page("ConfigPage")
        value = page_two.var2.get()
        for file in os.listdir(path_audio):
            filename = os.fsdecode(file)
            if filename.startswith(bird): 
                songs.append(filename)
        return songs

             
    def play_song(self,event):
         try:
            song= self.song.get()
            self.song.delete(0,"end")
            print(song)
            mixer.music.load(path_audio +(song))
            mixer.music.play(0)
            #a = mixer.Sound(path_audio +(song))
            #print("length of song is:",int(a.get_length()))    # return length of song in seconds
            #print("filename is: ", song_to_quiz)           # return filename of song
         except IndexError: 
            tk.messagebox.showerror(title="Show Error", message = "Oops! " + species_to_quiz +" audio files are missing")

app = BirdSong()
app.mainloop()